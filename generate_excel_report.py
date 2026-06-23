import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

history_file = Path("results/history.json")

if history_file.exists():

    with open(
        history_file,
        "r",
        encoding="utf-8"
    ) as f:

        history = json.load(f)

    history_df = pd.DataFrame(history)

else:

    history_df = pd.DataFrame(
        columns=[
            "time",
            "prediction",
            "confidence"
        ]
    )

metrics_file = Path(
    "results/metrics.csv"
)

if metrics_file.exists():

    metrics_df = pd.read_csv(
        metrics_file
    )

else:

    metrics_df = pd.DataFrame()

output_file = (
    "results/final_report.xlsx"
)

with pd.ExcelWriter(
    output_file,
    engine="openpyxl"
) as writer:

    metrics_df.to_excel(
        writer,
        sheet_name="Models",
        index=False
    )

    history_df.to_excel(
        writer,
        sheet_name="Predictions",
        index=False
    )

wb = load_workbook(
    output_file
)

for ws in wb.worksheets:

    for cell in ws[1]:
        cell.font = Font(
            bold=True
        )

    for column in ws.columns:

        max_len = 0

        column_letter = (
            get_column_letter(
                column[0].column
            )
        )

        for cell in column:

            try:

                if len(str(cell.value)) > max_len:
                    max_len = len(
                        str(cell.value)
                    )

            except:
                pass

        ws.column_dimensions[
            column_letter
        ].width = max_len + 3

wb.save(output_file)

print(
    f"Excel report saved: {output_file}"
)